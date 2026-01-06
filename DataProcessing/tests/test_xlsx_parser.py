import tempfile
import unittest
from collections.abc import Sequence
from decimal import Decimal
from pathlib import Path

try:
    import openpyxl
except Exception:  # pragma: no cover - optional dependency
    openpyxl = None

try:
    import pandas as pd
except Exception:  # pragma: no cover - optional dependency
    pd = None

from src.model.dataset_types import CupEquivalentUnit, PriceUnit
from src.model.series_code import get_series_code, normalize_form
from src.parsing.xlsx_parser import parse_xlsx


@unittest.skipIf(openpyxl is None or pd is None, "pandas/openpyxl required for XLSX tests")
class USDAFruitAndVegetablesXlsxParserTests(unittest.TestCase):
    def tearDown(self) -> None:
        if hasattr(self, "_temp_files"):
            for path in self._temp_files:
                try:
                    Path(path).unlink()
                except Exception:
                    pass

    def test_parses_xlsx_into_series_points(self) -> None:
        file_path = self._create_temp_xlsx(
            sheets=[
                (
                    "Apples",
                    "Apples - Average retail price per pound or pint and per cup equivalent, 2022",
                    [
                        [
                            "Form",
                            "Average retail price",
                            "",
                            "Preparation yield factor",
                            "Size of a cup equivalent",
                            "",
                            "Average price per cup equivalent",
                        ],
                        [
                            "Ready to drink3",
                            "0.869852963120289",
                            " per pint",
                            "1",
                            "0.5",
                            "Pints",
                            "0.4349264815601445",
                        ],
                    ],
                )
            ]
        )

        points = parse_xlsx(Path(file_path))
        self.assertEqual(len(points), 1)

        point = points[0]
        self.assertEqual(point.series_code, "apples_ready_to_drink")
        self.assertEqual(point.product_name, "Apples")
        self.assertEqual(point.form, "Ready to drink")
        self.assertEqual(point.date.year, 2022)

        self.assertEqual(point.average_retail_price, Decimal("0.869852963120289"))
        self.assertEqual(point.unit, PriceUnit.PER_PINT)
        self.assertEqual(point.preparation_yield_factor, Decimal("1"))
        self.assertEqual(point.cup_equivalent_size, Decimal("0.5"))
        self.assertEqual(point.cup_equivalent_unit, CupEquivalentUnit.PINTS)
        self.assertEqual(point.price_per_cup_equivalent, Decimal("0.4349264815601445"))

    def test_parses_split_header_rows(self) -> None:
        file_path = self._create_temp_xlsx(
            sheets=[
                (
                    "Apples",
                    "Apples - Average retail price per pound or pint and per cup equivalent, 2013",
                    [
                        [
                            "Form",
                            "Average retail price ",
                            "",
                            "Preparation",
                            "Size of a ",
                            "",
                            "Average price",
                        ],
                        [
                            "",
                            "",
                            "",
                            "yield factor",
                            "cup equivalent ",
                            "",
                            "per cup equivalent",
                        ],
                        [
                            "Ready to drink3",
                            "0.727287713104994",
                            " per pint",
                            "1",
                            "8",
                            "fl oz",
                            "0.363643856552497",
                        ],
                    ],
                )
            ]
        )

        points = parse_xlsx(Path(file_path))
        self.assertEqual(len(points), 1)
        point = points[0]
        self.assertEqual(point.series_code, "apples_ready_to_drink")
        self.assertEqual(point.cup_equivalent_unit, CupEquivalentUnit.FLUID_OUNCES)

    def test_parses_header_with_unit_columns(self) -> None:
        file_path = self._create_temp_xlsx(
            sheets=[
                (
                    "Acorn squash",
                    "Acorn squash - Average retail price per pound and per cup equivalent, 2023",
                    [
                        [
                            "Form",
                            "Average retail price ",
                            "Average retail price unit of measure",
                            "Preparation yield factor",
                            "Size of a cup equivalent",
                            "Cup equivalent unit of measure",
                            "Average price per cup equivalent ",
                        ],
                        ["Fresh1", "1.2414", " per pound", "0.45", "0.4519", "Pounds", "1.2235"],
                    ],
                )
            ]
        )

        points = parse_xlsx(Path(file_path))
        self.assertEqual(len(points), 1)
        self.assertEqual(points[0].series_code, "acorn_squash_fresh")

    def test_parses_annotated_price_unit(self) -> None:
        file_path = self._create_temp_xlsx(
            sheets=[
                (
                    "Apples",
                    "Apples - Average retail price per pound or pint and per cup equivalent, 2023",
                    [
                        [
                            "Form",
                            "Average retail price",
                            "",
                            "Preparation yield factor",
                            "Size of a cup equivalent",
                            "",
                            "Average price per cup equivalent",
                        ],
                        [
                            "Juice, ready to drink3",
                            "0.727287713104994",
                            "per pint (16 fluid ounces ready to drink)",
                            "1",
                            "8",
                            "fl oz",
                            "0.363643856552497",
                        ],
                    ],
                )
            ]
        )

        points = parse_xlsx(Path(file_path))
        self.assertEqual(len(points), 1)
        self.assertEqual(points[0].unit, PriceUnit.PER_PINT)

    def test_parses_singular_cup_unit(self) -> None:
        file_path = self._create_temp_xlsx(
            sheets=[
                (
                    "Kale",
                    "Kale - Average retail price per pound and per cup equivalent, 2013",
                    [
                        [
                            "Form",
                            "Average retail price",
                            "",
                            "Preparation yield factor",
                            "Size of a cup equivalent",
                            "",
                            "Average price per cup equivalent",
                        ],
                        ["Fresh1", "1.24", "per pound", "0.45", "0.45", "pound", "1.22"],
                    ],
                )
            ]
        )

        points = parse_xlsx(Path(file_path))
        self.assertEqual(len(points), 1)
        self.assertEqual(points[0].cup_equivalent_unit, CupEquivalentUnit.POUNDS)

    def test_raises_when_header_row_has_unexpected_unit_columns(self) -> None:
        file_path = self._create_temp_xlsx(
            sheets=[
                (
                    "Apples",
                    "Apples - Average retail price per pound or pint and per cup equivalent, 2022",
                    [
                        [
                            "Form",
                            "Average retail price",
                            "Unit",
                            "Preparation yield factor",
                            "Size of a cup equivalent",
                            "Cup equivalent unit",
                            "Average price per cup equivalent",
                        ],
                        ["Fresh1", "2", "per pound", "1", "0.25", "Pounds", "0.5"],
                    ],
                )
            ]
        )

        with self.assertRaises(ValueError):
            parse_xlsx(Path(file_path))

    def test_raises_when_price_unit_missing(self) -> None:
        file_path = self._create_temp_xlsx(
            sheets=[
                (
                    "Grapefruit",
                    "Grapefruit - Average retail price per pound or pint and per cup equivalent, 2022",
                    [
                        [
                            "Form",
                            "Average retail price",
                            "",
                            "Preparation yield factor",
                            "Size of a cup equivalent",
                            "",
                            "Average price per cup equivalent",
                        ],
                        [
                            "Juice, ready to drink2",
                            "1.19351651630181",
                            "",
                            "1",
                            "0.5",
                            "Pints",
                            "0.596758258150905",
                        ],
                    ],
                )
            ]
        )

        with self.assertRaises(ValueError):
            parse_xlsx(Path(file_path))

    def test_raises_when_cup_equivalent_unit_missing(self) -> None:
        file_path = self._create_temp_xlsx(
            sheets=[
                (
                    "Grapefruit",
                    "Grapefruit - Average retail price per pound or pint and per cup equivalent, 2022",
                    [
                        [
                            "Form",
                            "Average retail price",
                            "",
                            "Preparation yield factor",
                            "Size of a cup equivalent",
                            "",
                            "Average price per cup equivalent",
                        ],
                        [
                            "Juice, ready to drink2",
                            "1.19351651630181",
                            "per pint",
                            "1",
                            "0.5",
                            "",
                            "0.596758258150905",
                        ],
                    ],
                )
            ]
        )

        with self.assertRaises(ValueError):
            parse_xlsx(Path(file_path))

    def test_raises_when_price_unit_unknown(self) -> None:
        file_path = self._create_temp_xlsx(
            sheets=[
                (
                    "Apples",
                    "Apples - Average retail price per pound or pint and per cup equivalent, 2022",
                    [
                        [
                            "Form",
                            "Average retail price",
                            "",
                            "Preparation yield factor",
                            "Size of a cup equivalent",
                            "",
                            "Average price per cup equivalent",
                        ],
                        ["Fresh1", "2", "per lb", "1", "0.25", "Pounds", "0.5"],
                    ],
                )
            ]
        )

        with self.assertRaises(ValueError):
            parse_xlsx(Path(file_path))

    def test_raises_when_cup_equivalent_unit_unknown(self) -> None:
        file_path = self._create_temp_xlsx(
            sheets=[
                (
                    "Apples",
                    "Apples - Average retail price per pound or pint and per cup equivalent, 2022",
                    [
                        [
                            "Form",
                            "Average retail price",
                            "",
                            "Preparation yield factor",
                            "Size of a cup equivalent",
                            "",
                            "Average price per cup equivalent",
                        ],
                        ["Fresh1", "2", "per pound", "1", "0.25", "ounces", "0.5"],
                    ],
                )
            ]
        )

        with self.assertRaises(ValueError):
            parse_xlsx(Path(file_path))

    def test_raises_when_preparation_yield_factor_missing(self) -> None:
        file_path = self._create_temp_xlsx(
            sheets=[
                (
                    "Apples",
                    "Apples - Average retail price per pound or pint and per cup equivalent, 2022",
                    [
                        [
                            "Form",
                            "Average retail price",
                            "",
                            "Preparation yield factor",
                            "Size of a cup equivalent",
                            "",
                            "Average price per cup equivalent",
                        ],
                        [
                            "Ready to drink3",
                            "0.869852963120289",
                            "per pint",
                            "",
                            "0.5",
                            "Pints",
                            "0.4349264815601445",
                        ],
                    ],
                )
            ]
        )

        with self.assertRaises(ValueError):
            parse_xlsx(Path(file_path))

    def test_normalizes_series_code_slugs(self) -> None:
        self.assertEqual(get_series_code("Romaine lettuce", "Fresh1"), "romaine_lettuce_fresh")
        self.assertEqual(
            get_series_code("Cherries", "Canned, packed in syrup or water2"),
            "cherries_canned_packed_in_syrup_or_water",
        )
        self.assertEqual(normalize_form("Dried (Prunes)2"), "Dried (Prunes)")
        self.assertEqual(normalize_form("Juice, ready to drink2"), "Juice, Ready to drink")
        self.assertEqual(normalize_form("Fresh, Full Heads2"), "Fresh, Full heads")

    def test_applies_group_header_context_to_forms(self) -> None:
        file_path = self._create_temp_xlsx(
            sheets=[
                (
                    "Apples",
                    "Apples - Average retail price per pound or pint and per cup equivalent, 2022",
                    [
                        [
                            "Form",
                            "Average retail price",
                            "",
                            "Preparation yield factor",
                            "Size of a cup equivalent",
                            "",
                            "Average price per cup equivalent",
                        ],
                        ["Fresh1", "", "", "", "", "", ""],
                        [
                            "Ready to drink3",
                            "0.869852963120289",
                            " per pint",
                            "1",
                            "8",
                            "Fluid ounces",
                            "0.4349264815601445",
                        ],
                    ],
                )
            ]
        )

        points = parse_xlsx(Path(file_path))
        self.assertEqual(len(points), 1)
        self.assertEqual(points[0].series_code, "apples_fresh_ready_to_drink")
        self.assertEqual(points[0].form, "Fresh, Ready to drink")

    def test_raises_on_unknown_group_header(self) -> None:
        file_path = self._create_temp_xlsx(
            sheets=[
                (
                    "Apples",
                    "Apples - Average retail price per pound or pint and per cup equivalent, 2022",
                    [
                        [
                            "Form",
                            "Average retail price",
                            "",
                            "Preparation yield factor",
                            "Size of a cup equivalent",
                            "",
                            "Average price per cup equivalent",
                        ],
                        ["Frozen", "", "", "", "", "", ""],
                        [
                            "Ready to drink3",
                            "0.869852963120289",
                            " per pint",
                            "1",
                            "8",
                            "Fluid ounces",
                            "0.4349264815601445",
                        ],
                    ],
                )
            ]
        )

        with self.assertRaises(ValueError):
            parse_xlsx(Path(file_path))

    def test_accepts_mixed_vegetables_group_header(self) -> None:
        file_path = self._create_temp_xlsx(
            sheets=[
                (
                    "Mixed Vegetables",
                    "Mixed vegetables - Average retail price per pound and per cup equivalent, 2022",
                    [
                        [
                            "Form",
                            "Average retail price",
                            "",
                            "Preparation yield factor",
                            "Size of a cup equivalent",
                            "",
                            "Average price per cup equivalent",
                        ],
                        ["Peas & Carrots", "", "", "", "", "", ""],
                        [
                            "Canned1",
                            "1.475375",
                            "per pound",
                            "0.65",
                            "0.339173",
                            "Pounds",
                            "0.769857",
                        ],
                        [
                            "Frozen2",
                            "1.843287",
                            "per pound",
                            "0.99",
                            "0.339173",
                            "Pounds",
                            "0.631508",
                        ],
                    ],
                )
            ]
        )

        points = parse_xlsx(Path(file_path))
        self.assertEqual(len(points), 2)
        self.assertEqual(points[0].series_code, "mixed_vegetables_peas_carrots_canned")
        self.assertEqual(points[1].series_code, "mixed_vegetables_peas_carrots_frozen")

    def test_accepts_green_peas_and_carrots_group_header(self) -> None:
        file_path = self._create_temp_xlsx(
            sheets=[
                (
                    "Mixed Vegetables",
                    "Mixed vegetables - Average retail price per pound and per cup equivalent, 2013",
                    [
                        [
                            "Form",
                            "Average retail price",
                            "",
                            "Preparation yield factor",
                            "Size of a cup equivalent",
                            "",
                            "Average price per cup equivalent",
                        ],
                        ["Green Peas & Carrots", "", "", "", "", "", ""],
                        [
                            "Canned1",
                            "1.475375",
                            "per pound",
                            "0.65",
                            "0.339173",
                            "Pounds",
                            "0.769857",
                        ],
                        [
                            "Frozen2",
                            "1.843287",
                            "per pound",
                            "0.99",
                            "0.339173",
                            "Pounds",
                            "0.631508",
                        ],
                    ],
                )
            ]
        )

        points = parse_xlsx(Path(file_path))
        self.assertEqual(len(points), 2)
        self.assertEqual(points[0].series_code, "mixed_vegetables_green_peas_carrots_canned")
        self.assertEqual(points[1].series_code, "mixed_vegetables_green_peas_carrots_frozen")

    def test_accepts_succotash_group_header(self) -> None:
        file_path = self._create_temp_xlsx(
            sheets=[
                (
                    "Mixed Vegetables",
                    "Mixed vegetables - Average retail price per pound and per cup equivalent, 2013",
                    [
                        [
                            "Form",
                            "Average retail price",
                            "",
                            "Preparation yield factor",
                            "Size of a cup equivalent",
                            "",
                            "Average price per cup equivalent",
                        ],
                        ["Succotash3", "", "", "", "", "", ""],
                        [
                            "Canned1",
                            "1.475375",
                            "per pound",
                            "0.65",
                            "0.339173",
                            "Pounds",
                            "0.769857",
                        ],
                        [
                            "Frozen2",
                            "1.843287",
                            "per pound",
                            "0.99",
                            "0.339173",
                            "Pounds",
                            "0.631508",
                        ],
                    ],
                )
            ]
        )

        points = parse_xlsx(Path(file_path))
        self.assertEqual(len(points), 2)
        self.assertEqual(points[0].series_code, "mixed_vegetables_succotash_canned")
        self.assertEqual(points[1].series_code, "mixed_vegetables_succotash_frozen")

    def test_raises_on_missing_form_with_other_data(self) -> None:
        file_path = self._create_temp_xlsx(
            sheets=[
                (
                    "Apples",
                    "Apples - Average retail price per pound or pint and per cup equivalent, 2022",
                    [
                        [
                            "Form",
                            "Average retail price",
                            "",
                            "Preparation yield factor",
                            "Size of a cup equivalent",
                            "",
                            "Average price per cup equivalent",
                        ],
                        ["", "0.25", "per pound", "1", "0.5", "Pounds", "0.5"],
                    ],
                )
            ]
        )

        with self.assertRaises(ValueError):
            parse_xlsx(Path(file_path))

    def test_skips_footnote_rows(self) -> None:
        file_path = self._create_temp_xlsx(
            sheets=[
                (
                    "Acorn squash",
                    "Acorn squash - Average retail price per pound or pint and per cup equivalent, 2023",
                    [
                        [
                            "Form",
                            "Average retail price",
                            "",
                            "Preparation yield factor",
                            "Size of a cup equivalent",
                            "",
                            "Average price per cup equivalent",
                        ],
                        ["Fresh1", "2", "per pound", "1", "0.25", "Pounds", "0.5"],
                        [
                            "1It is assumed that consumers bake acorn squash prior to consumption.",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                        ],
                        [
                            "Source: USDA, Economic Research Service (ERS) calculations.",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                        ],
                        [
                            "Contact: Hayden Stewart or Jeffrey Hyman.",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                        ],
                        [
                            "Errata: On June 25, 2018, ERS revised this table.",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                        ],
                    ],
                )
            ]
        )

        points = parse_xlsx(Path(file_path))
        self.assertEqual(len(points), 1)
        self.assertEqual(points[0].series_code, "acorn_squash_fresh")

    def test_parses_all_sheets_in_workbook(self) -> None:
        file_path = self._create_temp_xlsx(
            sheets=[
                (
                    "Apples",
                    "Apples - Average retail price per pound or pint and per cup equivalent, 2022",
                    [
                        [
                            "Form",
                            "Average retail price",
                            "",
                            "Preparation yield factor",
                            "Size of a cup equivalent",
                            "",
                            "Average price per cup equivalent",
                        ],
                        ["Fresh1", "2", "per pound", "1", "0.25", "Pounds", "0.5"],
                    ],
                ),
                (
                    "Grapefruit",
                    "Grapefruit - Average retail price per pound or pint and per cup equivalent, 2022",
                    [
                        [
                            "Form",
                            "Average retail price",
                            "",
                            "Preparation yield factor",
                            "Size of a cup equivalent",
                            "",
                            "Average price per cup equivalent",
                        ],
                        [
                            "Juice, ready to drink2",
                            "1.19351651630181",
                            "per pint",
                            "1",
                            "8",
                            "Fluid ounces",
                            "0.596758258150905",
                        ],
                    ],
                ),
            ]
        )

        points = parse_xlsx(Path(file_path))
        self.assertEqual(len(points), 2)
        self.assertEqual(points[0].series_code, "apples_fresh")
        self.assertEqual(points[1].series_code, "grapefruit_juice_ready_to_drink")

    def test_raises_when_header_row_is_missing(self) -> None:
        file_path = self._create_temp_xlsx(
            sheets=[
                (
                    "Apples",
                    "Apples - Average retail price per pound or pint and per cup equivalent, 2022",
                    [["Not a header", "Still not", "", "", "", "", ""]],
                )
            ]
        )

        with self.assertRaises(ValueError):
            parse_xlsx(Path(file_path))

    def test_raises_when_header_row_has_mismatched_columns(self) -> None:
        file_path = self._create_temp_xlsx(
            sheets=[
                (
                    "Apples",
                    "Apples - Average retail price per pound or pint and per cup equivalent, 2022",
                    [
                        [
                            "Form",
                            "Average retail price",
                            "",
                            "Yield factor",
                            "Size of a cup equivalent",
                            "",
                            "Average price per cup equivalent",
                        ],
                        ["Fresh1", "2", "per pound", "1", "0.25", "Pounds", "0.5"],
                    ],
                )
            ]
        )

        with self.assertRaises(ValueError):
            parse_xlsx(Path(file_path))

    def test_raises_when_sheet_title_has_no_year(self) -> None:
        """Per Constitution: Fail-fast - sheets must contain 4-digit year."""
        file_path = self._create_temp_xlsx(
            sheets=[
                (
                    "Apples",
                    "Apples - Average retail price per pound or pint and per cup equivalent",  # no year
                    [
                        [
                            "Form",
                            "Average retail price",
                            "",
                            "Preparation yield factor",
                            "Size of a cup equivalent",
                            "",
                            "Average price per cup equivalent",
                        ],
                        ["Fresh1", "2", "per pound", "1", "0.25", "Pounds", "0.5"],
                    ],
                )
            ]
        )

        with self.assertRaises(ValueError) as context:
            parse_xlsx(Path(file_path))
        self.assertIn("year", str(context.exception).lower())

    def _create_temp_xlsx(self, sheets: Sequence[tuple[str, str, Sequence[Sequence[str | None]]]]) -> str:
        if not hasattr(self, "_temp_files"):
            self._temp_files: list[str] = []
        assert openpyxl is not None
        wb = openpyxl.Workbook()

        for index, (sheet_name, title, rows) in enumerate(sheets):
            if index == 0:
                ws = wb.active
                if ws is None:
                    raise RuntimeError("Workbook missing active sheet.")
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)
            assert ws is not None

            ws.cell(row=1, column=1, value=title)
            row_index = 2
            for row in rows:
                for col_index, value in enumerate(row, start=1):
                    ws.cell(row=row_index, column=col_index, value=value)
                row_index += 1

        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        temp_file.close()
        wb.save(temp_file.name)
        self._temp_files.append(temp_file.name)
        return temp_file.name


if __name__ == "__main__":
    unittest.main()
